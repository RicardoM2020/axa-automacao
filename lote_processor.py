"""
Processamento do relatório ERP → Lote AXA.
Aplica todas as regras de negócio, validações e controle de duplo faturamento.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import gspread
from google.oauth2.service_account import Credentials
import warnings
warnings.filterwarnings("ignore")

from config import (
    SHEET_ID, ABA_HIST_FATURADOS, ABA_MATRIZ,
    OUTPUT_DIR, LOTE_ATUAL, COBERTURAS_NORMALIZAR,
    FINALIZACOES_COM_DESPESA, FINALIZACOES_SEM_DESPESA,
)
from matriz_parametros import calcular_valores

# Escopo do Google Sheets
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# Caminho para credenciais da service account (opcional)
CREDENTIALS_FILE = Path(__file__).parent / "google_credentials.json"


# ---------------------------------------------------------------------------
# Leitura de dados do Google Sheets
# ---------------------------------------------------------------------------

def _conectar_sheets():
    """Conecta ao Google Sheets via OAuth. Reutiliza token salvo quando possível."""
    import pickle
    from google.auth.transport.requests import Request
    from google_auth_oauthlib.flow import InstalledAppFlow

    token_path = Path(__file__).parent / "token.pickle"
    oauth_path = Path(__file__).parent / "oauth_credentials.json"
    creds = None

    if token_path.exists():
        with open(token_path, "rb") as f:
            creds = pickle.load(f)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(oauth_path), SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, "wb") as f:
            pickle.dump(creds, f)

    return gspread.authorize(creds)


def carregar_hist_faturados() -> pd.DataFrame:
    """Carrega o histórico de sinistros já faturados do Google Sheets."""
    print("[Sheets] Carregando Hist. Faturados...")
    gc = _conectar_sheets()
    sh = gc.open_by_key(SHEET_ID)
    aba = sh.worksheet(ABA_HIST_FATURADOS)

    # Lê todos os valores brutos para evitar erro de cabeçalhos duplicados
    valores = aba.get_all_values()
    if not valores:
        return pd.DataFrame()

    cabecalho = valores[0]
    # Remove colunas vazias e resolve duplicatas
    cols_vistas = {}
    cabecalho_limpo = []
    for c in cabecalho:
        c = c.strip()
        if not c:
            c = f"_col_{len(cabecalho_limpo)}"
        if c in cols_vistas:
            cols_vistas[c] += 1
            c = f"{c}_{cols_vistas[c]}"
        else:
            cols_vistas[c] = 0
        cabecalho_limpo.append(c)

    df = pd.DataFrame(valores[1:], columns=cabecalho_limpo)
    # Remove colunas geradas automaticamente (_col_X)
    df = df.loc[:, ~df.columns.str.startswith("_col_")]
    print(f"[Sheets] {len(df)} registros no historico.")
    return df


def salvar_hist_faturados(novos_registros: pd.DataFrame):
    """Adiciona os sinistros do novo lote ao histórico."""
    print("[Sheets] Atualizando Hist. Faturados...")
    gc = _conectar_sheets()
    sh = gc.open_by_key(SHEET_ID)
    aba = sh.worksheet(ABA_HIST_FATURADOS)

    # Converte para lista de linhas
    linhas = novos_registros.fillna("").values.tolist()
    aba.append_rows(linhas, value_input_option="USER_ENTERED")
    print(f"[Sheets] {len(linhas)} registros adicionados ao histórico.")


# ---------------------------------------------------------------------------
# Limpeza e normalização do relatório ERP
# ---------------------------------------------------------------------------

def limpar_relatorio(df: pd.DataFrame) -> pd.DataFrame:
    """Aplica todas as limpezas necessárias no arquivo exportado do ERP."""
    print("[Limpeza] Iniciando limpeza do relatório...")
    df = df.copy()

    # Remove linhas completamente vazias
    df = df.dropna(how="all")

    # Strip de espaços e quebras de linha em todas as colunas de texto
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip().str.replace("\r\n", " ", regex=False).str.replace("\r", " ", regex=False).str.replace("\n", " ", regex=False).str.strip()

    # Limpa valores monetários (remove R$, espaços, troca ponto por vírgula → float)
    colunas_monetarias = [
        "Valor Pagto Indenização", "Valor Pagto Mão-de-obra",
        "Valor Pagto Honorário", "Valor Pagto Despesa",
        "LMI DA APÓLICE", "VALOR FRANQUIA",
        "reserva_financeira_aprovacao", "reserva_despesa_valor",
        "reserva_indenizacao_valor",
    ]
    for col in colunas_monetarias:
        if col in df.columns:
            def _para_float(v):
                s = str(v).strip().replace("R$", "").replace("\r", "").replace("\n", "").strip()
                if not s or s in ("-", "nan", "None"):
                    return 0.0
                # Formato brasileiro: 1.234,56 → remove ponto, troca vírgula
                if "," in s and "." in s:
                    s = s.replace(".", "").replace(",", ".")
                elif "," in s:
                    s = s.replace(",", ".")
                # Se só tem ponto (ex: 449.00), já está no formato correto
                try:
                    return float(s)
                except ValueError:
                    return 0.0
            df[col] = df[col].apply(_para_float)

    # Remove apóstrofo do número de sinistro e CPF
    for col in ["Sinistro AXA", "CPF", "CPF/CNPJ Beneficiário"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.lstrip("'").str.strip()

    # Normaliza coberturas (aceita tanto "Cobertura de Seguro" quanto "cobertura")
    col_cob = next((c for c in df.columns if c.lower().strip() in ("cobertura de seguro", "cobertura")), None)
    if col_cob and col_cob != "Cobertura de Seguro":
        df = df.rename(columns={col_cob: "Cobertura de Seguro"})
    if "Cobertura de Seguro" in df.columns:
        df["Cobertura de Seguro"] = df["Cobertura de Seguro"].astype(str).str.strip()
        for cobertura_antiga in COBERTURAS_NORMALIZAR:
            mask = df["Cobertura de Seguro"].str.lower() == cobertura_antiga.lower()
            df.loc[mask, "Cobertura de Seguro"] = "Quebra Acidental"
            if mask.sum() > 0:
                print(f"[Limpeza] '{cobertura_antiga}' -> 'Quebra Acidental': {mask.sum()} linha(s)")

        # Reclassifica Quebra Acidental com franquia = 0 para Extensão da Garantia Original
        col_pct_f = next((c for c in df.columns if "PERCENTUAL" in c.upper()), None)
        if col_pct_f:
            mask_qa_zero = (
                (df["Cobertura de Seguro"] == "Quebra Acidental") &
                (df[col_pct_f].fillna(0).astype(float) == 0)
            )
            if mask_qa_zero.sum() > 0:
                df.loc[mask_qa_zero, "Cobertura de Seguro"] = "Extensão da Garantia Original"
                print(f"[Limpeza] {int(mask_qa_zero.sum())} linha(s) Quebra Acidental (franquia=0) -> Extensao da Garantia Original.")

        # Preenche cobertura vazia usando a coluna NATUREZA
        NATUREZA_PARA_COBERTURA = {
            "quebra acidental": "Quebra Acidental",
            "dano funcional":   "Quebra Acidental",
            "furto":            "Roubo e Furto",
            "roubo":            "Roubo e Furto",
            "dano eletrico":    "Quebra Acidental",
            "danos eletricos":  "Quebra Acidental",
            "derramamento":     "Quebra Acidental",
        }
        col_nat = next((c for c in df.columns if c.upper() == "NATUREZA"), None)

        def _eh_vazia(serie):
            return serie.isna() | serie.astype(str).str.strip().isin(["", "nan", "None"])

        if col_nat:
            mask_vazia = _eh_vazia(df["Cobertura de Seguro"])
            preenchidos = 0
            for nat_key, cob_valor in NATUREZA_PARA_COBERTURA.items():
                mask_nat = mask_vazia & df[col_nat].astype(str).str.lower().str.strip().str.startswith(nat_key)
                df.loc[mask_nat, "Cobertura de Seguro"] = cob_valor
                preenchidos += int(mask_nat.sum())
                mask_vazia = _eh_vazia(df["Cobertura de Seguro"])
            if preenchidos > 0:
                print(f"[Limpeza] {preenchidos} linha(s) com cobertura preenchida via coluna NATUREZA.")
        # Fallback: usa DESCRIÇÃO DO EVENTO/RECLAMAÇÃO quando NATUREZA também está vazia
        DESCRICAO_PARA_COBERTURA = {
            "roubado":      "Roubo e Furto",
            "roubo":        "Roubo e Furto",
            "furtado":      "Roubo e Furto",
            "furto":        "Roubo e Furto",
            "subtra":       "Roubo e Furto",
            "arrombamento": "Roubo e Furto",
            "quebr":        "Quebra Acidental",
            "queda":        "Quebra Acidental",
            "dano":         "Quebra Acidental",
            "molh":         "Quebra Acidental",
            "liquid":       "Quebra Acidental",
            "agua":         "Quebra Acidental",
            "garantia":     "Extensão da Garantia Original",
        }
        col_desc = next((c for c in df.columns if "DESCRI" in c.upper() and "EVENTO" in c.upper()), None)
        if col_desc:
            mask_ainda = _eh_vazia(df["Cobertura de Seguro"])
            preenchidos_desc = 0
            for palavra, cob_valor in DESCRICAO_PARA_COBERTURA.items():
                mask_desc = mask_ainda & df[col_desc].astype(str).str.lower().str.contains(palavra, na=False)
                df.loc[mask_desc, "Cobertura de Seguro"] = cob_valor
                preenchidos_desc += int(mask_desc.sum())
                mask_ainda = _eh_vazia(df["Cobertura de Seguro"])
            if preenchidos_desc > 0:
                print(f"[Limpeza] {preenchidos_desc} linha(s) com cobertura preenchida via DESCRICAO DO EVENTO.")

        ainda_vazias = int(_eh_vazia(df["Cobertura de Seguro"]).sum())
        if ainda_vazias > 0:
            print(f"[Limpeza] Aviso: {ainda_vazias} linha(s) ainda sem cobertura — serao excluidas.")

    # Normaliza Tipo de Aparelho para maiúsculo
    if "Tipo de Aparelho" in df.columns:
        df["Tipo de Aparelho"] = df["Tipo de Aparelho"].astype(str).str.upper().str.strip()
        # Vazios → CELULAR
        mask_vazio_ap = df["Tipo de Aparelho"].isna() | df["Tipo de Aparelho"].isin(["NAN", ""])
        if mask_vazio_ap.sum() > 0:
            df.loc[mask_vazio_ap, "Tipo de Aparelho"] = "CELULAR"
            print(f"[Limpeza] {int(mask_vazio_ap.sum())} linha(s) sem tipo de aparelho -> CELULAR.")

        # Tudo que não é CELULAR nem TABLET → TABLET (inclui NOTEBOOK, TELEFONE FIXO, etc.)
        mask_outros = ~df["Tipo de Aparelho"].isin(["CELULAR", "TABLET"])
        if mask_outros.sum() > 0:
            outros = df.loc[mask_outros, "Tipo de Aparelho"].unique().tolist()
            df.loc[mask_outros, "Tipo de Aparelho"] = "TABLET"
            print(f"[Limpeza] {int(mask_outros.sum())} linha(s) {outros} -> TABLET.")

        # Corrige CELULAR para TABLET quando Valor Pagto Despesa do ERP é 127,07
        # (diferença de -38,12 entre despesa calculada 88,95 e despesa ERP 127,07)
        col_vl_desp_raw = _encontrar_coluna_fuzzy(df, ["Valor Pagto Despesa"])
        if col_vl_desp_raw:
            def _to_float_desp(v):
                s = str(v).strip().replace("R$","").replace("\r","").replace("\n","").strip()
                if not s or s in ("-","nan","None"): return 0.0
                if "," in s and "." in s: s = s.replace(".","").replace(",",".")
                elif "," in s: s = s.replace(",",".")
                try: return float(s)
                except: return 0.0
            desp_erp = df[col_vl_desp_raw].apply(_to_float_desp)
            mask_tablet = (df["Tipo de Aparelho"] == "CELULAR") & (desp_erp.round(2) == 127.07)
            if mask_tablet.sum() > 0:
                df.loc[mask_tablet, "Tipo de Aparelho"] = "TABLET"
                print(f"[Limpeza] {int(mask_tablet.sum())} linha(s) corrigidas CELULAR -> TABLET (despesa ERP = 127,07).")

    # Preenche Status/Finalização com base na coluna Observações quando estão vazios
    col_obs = _encontrar_col_simples(df, ["Observações", "Observacoes", "Observaçoes"])
    col_status_l = _encontrar_col_simples(df, ["Status do Sinistro"])
    col_fin_l    = _encontrar_col_simples(df, ["Tipo de Finalização", "Tipo de Finalizacao"])

    col_setor = next((c for c in df.columns if c.lower() == "setor_status"), None)

    if col_status_l and col_fin_l:
        def _status_vazio(df):
            return df[col_status_l].isna() | df[col_status_l].astype(str).str.strip().isin(["", "nan"])

        # Via coluna Observações: "NEGADO - ..."
        if col_obs:
            mask_vazio = _status_vazio(df)
            obs_upper = df.loc[mask_vazio, col_obs].astype(str).str.upper()
            mask_negado = mask_vazio & obs_upper.str.contains("NEGADO", na=False)
            df.loc[mask_negado, col_status_l] = "Negado"
            df.loc[mask_negado, col_fin_l]    = "Negado"
            if mask_negado.sum() > 0:
                print(f"[Limpeza] {int(mask_negado.sum())} linha(s) preenchidas como Negado/Negado via Observacoes.")

        # Via setor_status: "Finalizado Sem Atendimento", "Finalizado Indenizado", etc.
        if col_setor:
            SETOR_MAP = {
                "sem atendimento": ("Encerrado", "Sem Atendimento"),
                "indenizado":      ("Indenizado", "Crédito em Conta"),
                "negado":          ("Negado",     "Negado"),
                "reparo":          ("Indenizado", "Reparo"),
                "reposicao":       ("Indenizado", "Reposição"),
                "reposi":          ("Indenizado", "Reposição"),
            }
            mask_vazio = _status_vazio(df)
            setor_lower = df.loc[mask_vazio, col_setor].astype(str).str.lower()
            for palavra, (status_val, fin_val) in SETOR_MAP.items():
                mask_s = mask_vazio & setor_lower.str.contains(palavra, na=False)
                df.loc[mask_s, col_status_l] = status_val
                df.loc[mask_s, col_fin_l]    = fin_val
                mask_vazio = _status_vazio(df)
                setor_lower = df.loc[mask_vazio, col_setor].astype(str).str.lower() if mask_vazio.sum() > 0 else setor_lower
            preenchidos_setor = (~_status_vazio(df)).sum() - (~_status_vazio(df)).sum()
            # Conta quantos foram preenchidos via setor
            total_preench = int((~_status_vazio(df)).sum())

        # Ainda vazios
        ainda_vazios = int(_status_vazio(df).sum())
        if ainda_vazios > 0:
            print(f"[Limpeza] Aviso: {ainda_vazios} linha(s) com Status vazio — serao excluidas do lote.")

    print(f"[Limpeza] Concluída. {len(df)} linhas válidas.")
    return df


# ---------------------------------------------------------------------------
# Verificação da estrutura de colunas
# ---------------------------------------------------------------------------

COLUNAS_ESPERADAS = [
    "Sinistro AXA", "Nome do cliente", "CPF", "Ordem de Serviço",
    "CPF/CNPJ Beneficiário", "Número Lote", "Razão", "Cobertura de Seguro",
    "Valor Pagto Indenização", "Valor Pagto Mão-de-obra", "Valor Pagto Honorário",
    "Valor Pagto Despesa", "Invoice/Nota", "Service Code",
    "Data do Recebimento do Equipamento/Documento", "Data Orçamento",
    "Data Conclusão do Serviço", "Produto/Laudo entregue",
    "Status do Sinistro", "Tipo de Finalização", "Motivo de Troca",
    "Observações", "Parecer", "Data de Conclusão",
    "LMI DA APÓLICE", "PARCEIRO", "Percentual", "Tipo de Aparelho",
    "OS Interna PLL", "Unidade PLL",
]


def verificar_colunas(df: pd.DataFrame) -> list[str]:
    """Retorna lista de colunas esperadas que estão faltando no arquivo ERP."""
    faltando = [c for c in COLUNAS_ESPERADAS if c not in df.columns]
    if faltando:
        print(f"[Estrutura] Aviso: {len(faltando)} coluna(s) não encontrada(s): {faltando}")
    else:
        print("[Estrutura] Todas as colunas esperadas estão presentes.")
    return faltando


# ---------------------------------------------------------------------------
# Controle de duplo faturamento
# ---------------------------------------------------------------------------

def marcar_faturamentos(df: pd.DataFrame, hist: pd.DataFrame) -> pd.DataFrame:
    """
    Cruza com o histórico e marca separadamente:
    - indenizacao_ja_faturada: OS tem registro com Valor Pagto Indenização > 0
    - despesa_ja_faturada:     OS tem registro com Valor Pagto Despesa > 0
    """
    df = df.copy()

    if hist.empty:
        df["indenizacao_ja_faturada"] = False
        df["despesa_ja_faturada"] = False
        return df

    col_os_hist   = _encontrar_coluna_fuzzy(hist, ["OS Interna PLL", "OS Interna"])
    col_os_df     = _encontrar_coluna_fuzzy(df,   ["OS Interna PLL", "OS Interna"])
    col_inden_hist = _encontrar_coluna_fuzzy(hist, ["Valor Pagto Indeniza"])
    col_desp_hist  = _encontrar_coluna_fuzzy(hist, ["Valor Pagto Despesa"])

    if not col_os_hist or not col_os_df:
        print("[Duplo Fat.] Aviso: coluna OS nao encontrada — verificacao ignorada.")
        df["indenizacao_ja_faturada"] = False
        df["despesa_ja_faturada"] = False
        return df

    def _val(s):
        v = str(s).replace("R$","").replace(".","").replace(",",".").strip()
        try: return float(v)
        except: return 0.0

    hist = hist.copy()
    hist["_os"]    = hist[col_os_hist].astype(str).str.strip()
    hist["_inden"] = hist[col_inden_hist].apply(_val) if col_inden_hist else 0
    hist["_desp"]  = hist[col_desp_hist].apply(_val)  if col_desp_hist  else 0

    # OS que têm indenização já faturada (valor > 0)
    os_inden_faturada = set(hist.loc[hist["_inden"] > 0.01, "_os"])
    # OS que têm despesa já faturada (valor > 0)
    os_desp_faturada  = set(hist.loc[hist["_desp"]  > 0.01, "_os"])

    os_df = df[col_os_df].astype(str).str.strip()
    df["indenizacao_ja_faturada"] = os_df.isin(os_inden_faturada)
    df["despesa_ja_faturada"]     = os_df.isin(os_desp_faturada)

    print(f"[Duplo Fat.] {int(df['indenizacao_ja_faturada'].sum())} OS ja com indenizacao faturada.")
    print(f"[Duplo Fat.] {int(df['despesa_ja_faturada'].sum())} OS ja com despesa faturada.")
    return df


def _encontrar_col_simples(df: pd.DataFrame, candidatos: list[str]) -> str | None:
    """Busca coluna por nome exato (case-insensitive)."""
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidatos:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    return None


def _encontrar_coluna(df: pd.DataFrame, candidatos: list[str]) -> str | None:
    for c in candidatos:
        if c in df.columns:
            return c
    return None


def _encontrar_coluna_fuzzy(df: pd.DataFrame, candidatos: list[str]) -> str | None:
    """Busca coluna por nome exato ou por prefixo (case-insensitive, ignora problemas de encoding)."""
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidatos:
        cand_lower = cand.lower()
        # Exato
        if cand_lower in cols_lower:
            return cols_lower[cand_lower]
        # Prefixo
        for col_l, col_orig in cols_lower.items():
            if col_l.startswith(cand_lower[:10]):
                return col_orig
    return None


# ---------------------------------------------------------------------------
# Aplicação das regras de faturamento
# ---------------------------------------------------------------------------

def aplicar_regras_faturamento(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aplica todas as regras de negócio e retorna o DataFrame com colunas calculadas:
    - valor_despesa_correto
    - pct_indenizacao
    - valor_indenizacao_correto
    - diferenca_despesa
    - diferenca_indenizacao
    - incluir_indenizacao (bool)
    - incluir_despesa (bool)
    - motivo_exclusao (str)
    - regra_aplicada (str)
    """
    df = df.copy()
    registros = []

    col_finalizacao = _encontrar_coluna_fuzzy(df, ["Tipo de Finalização", "Tipo de Finalizacao", "Tipo de Finaliza"])
    col_status      = _encontrar_coluna_fuzzy(df, ["Status do Sinistro", "Status"])
    col_cobertura   = _encontrar_coluna_fuzzy(df, ["Cobertura de Seguro", "cobertura"])
    col_aparelho    = _encontrar_coluna_fuzzy(df, ["Tipo de Aparelho"])
    col_lmi         = _encontrar_coluna_fuzzy(df, ["LMI DA AP", "LMI", "lmi_da_apolice"])
    col_vl_inden    = _encontrar_coluna_fuzzy(df, ["Valor Pagto Indeniza", "Valor Pagto Indenização"])
    col_vl_desp     = _encontrar_coluna_fuzzy(df, ["Valor Pagto Despesa"])
    col_estado_ap   = _encontrar_coluna_fuzzy(df, ["estado_aparelho", "Estado Aparelho"])
    col_pct_franquia= _encontrar_coluna_fuzzy(df, ["PERCENTUAL COBRADO", "pct_franquia", "Percentual"])
    col_parceiro    = _encontrar_coluna_fuzzy(df, ["PARCEIRO"])

    PARCEIROS_REPARO_ACEITAR_ERP = {
        "solinformatica", "infostore", "sol informatica",
        "sol informática", "papel & cia", "papel e cia", "papel",
    }

    for idx, row in df.iterrows():
        cobertura      = str(row.get(col_cobertura, "")).strip()
        status         = str(row.get(col_status, "")).strip()
        finalizacao    = str(row.get(col_finalizacao, "")).strip()
        aparelho       = str(row.get(col_aparelho, "CELULAR")).strip().upper()
        lmi            = float(row.get(col_lmi, 0) or 0) if col_lmi else 0
        vl_inden_erp   = float(row.get(col_vl_inden, 0) or 0) if col_vl_inden else 0
        vl_desp_erp    = float(row.get(col_vl_desp, 0) or 0) if col_vl_desp else 0
        estado_aparelho= str(row.get(col_estado_ap, "")).strip() if col_estado_ap else ""
        pct_franquia   = float(row.get(col_pct_franquia, 0) or 0) if col_pct_franquia else 0
        parceiro       = str(row.get(col_parceiro, "")).strip().lower() if col_parceiro else ""
        inden_faturada = bool(row.get("indenizacao_ja_faturada", False))
        desp_faturada  = bool(row.get("despesa_ja_faturada", False))

        # Calcula valores corretos pela matriz (fórmula: ((% inden - % franquia) × LMI) / 100)
        calc = calcular_valores(cobertura, status, finalizacao, aparelho, lmi, estado_aparelho, pct_franquia)

        incluir_inden = True
        incluir_desp  = True
        motivo        = ""

        # Regra 0: Encerrado/Sem Atendimento só fatura em Roubo e Furto
        if status == "Encerrado" and finalizacao == "Sem Atendimento" and cobertura not in ("Roubo e Furto",):
            incluir_inden = False
            incluir_desp  = False
            motivo = f"Sem Atendimento em '{cobertura}' — so fatura em Roubo e Furto."

        # Regra 0b: Dados insuficientes
        elif not cobertura or cobertura.lower() in ("nan", "none", ""):
            incluir_inden = False
            incluir_desp  = False
            motivo = "Cobertura vazia — sem dados suficientes."
        elif not status or status == "nan":
            incluir_inden = False
            incluir_desp  = False
            motivo = "Status vazio apos tentativa de preenchimento automatico."
        elif aparelho not in ("CELULAR", "TABLET", "NOTEBOOK"):
            incluir_inden = False
            incluir_desp  = False
            motivo = f"Tipo de Aparelho invalido ({aparelho}) — excluido."

        # Regra 1: Indenização já faturada → exclui indenização
        if inden_faturada:
            incluir_inden = False
            motivo += "Indenização já faturada. "

        # Regra 2: Despesa já faturada + finalização Negado/Sem Atendimento → exclui tudo
        if desp_faturada and finalizacao in FINALIZACOES_SEM_DESPESA:
            incluir_inden = False
            incluir_desp  = False
            motivo += "Despesa já faturada + finalização sem indenização. "

        # Regra 3: Despesa já faturada + finalização com indenização → zera despesa, mantém indenização
        elif desp_faturada and finalizacao in FINALIZACOES_COM_DESPESA:
            incluir_desp = False
            motivo += "Despesa já faturada anteriormente — cobrar apenas indenização. "

        # Regra 4: Se % indenização = 0, não há indenização a cobrar
        if calc["pct_indenizacao"] == 0 and not inden_faturada:
            incluir_inden = False

        # Regra 5: Reparo de parceiro homologado → aceita valor de indenização do ERP
        parceiro_homologado = any(p in parceiro for p in PARCEIROS_REPARO_ACEITAR_ERP)
        usar_valor_erp_inden = (
            finalizacao == "Reparo" and
            parceiro_homologado and
            incluir_inden and
            abs(calc["valor_indenizacao_correto"] - vl_inden_erp) > 0.01
        )
        valor_inden_final = vl_inden_erp if usar_valor_erp_inden else calc["valor_indenizacao_correto"]

        registros.append({
            "valor_despesa_correto":    calc["valor_despesa"] if incluir_desp else 0,
            "pct_indenizacao":          calc["pct_indenizacao"],
            "valor_indenizacao_correto":valor_inden_final if incluir_inden else 0,
            "diferenca_despesa":        round((calc["valor_despesa"] - vl_desp_erp), 2),
            "diferenca_indenizacao":    round((valor_inden_final - vl_inden_erp), 2),
            "incluir_indenizacao":      incluir_inden,
            "incluir_despesa":          incluir_desp,
            "usar_valor_erp_inden":     usar_valor_erp_inden,
            "motivo_exclusao":          motivo.strip(),
            "regra_aplicada":           calc["regra_aplicada"],
        })

    df_calc = pd.DataFrame(registros, index=df.index)
    return pd.concat([df, df_calc], axis=1)


# ---------------------------------------------------------------------------
# Geração dos outputs
# ---------------------------------------------------------------------------

def gerar_lote(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filtra apenas as linhas que têm algo a faturar e substitui os valores
    do ERP pelos valores recalculados (% × LMI) antes de exportar.
    """
    mask = df["incluir_indenizacao"] | df["incluir_despesa"]
    lote = df[mask].copy()

    # Substitui valor de indenização pelo recalculado
    col_vl_inden = _encontrar_coluna_fuzzy(lote, ["Valor Pagto Indeniza"])
    if col_vl_inden:
        lote[col_vl_inden] = lote["valor_indenizacao_correto"]

    # Substitui valor de despesa pelo recalculado
    col_vl_desp = _encontrar_coluna_fuzzy(lote, ["Valor Pagto Despesa"])
    if col_vl_desp:
        lote[col_vl_desp] = lote["valor_despesa_correto"]

    # Preenche Número Lote e Razão com o lote atual
    from config import LOTE_NUMERO
    CANDIDATOS_NUM_LOTE = {"número lote", "numero lote", "n\xfamero lote"}
    CANDIDATOS_RAZAO    = {"razão", "razao", "raz\xe3o"}
    for col in lote.columns:
        if col.strip().lower() in CANDIDATOS_NUM_LOTE:
            lote[col] = LOTE_NUMERO
        if col.strip().lower() in CANDIDATOS_RAZAO:
            lote[col] = "B2B 2.0"

    print(f"[Lote] {len(lote)} sinistros no lote (de {len(df)} processados).")
    return lote


def exportar_excel(df_lote: pd.DataFrame, df_completo: pd.DataFrame, df_hist: pd.DataFrame = None):
    """
    Gera dois arquivos Excel na pasta output/:
    1. Lote_AXA_<data>.xlsx — arquivo para enviar à AXA
    2. Relatorio_Conferencia_<data>.xlsx — relatório interno com todas as validações
    """
    data_hoje = datetime.today().strftime("%Y%m%d_%H%M")

    # --- Arquivo do Lote ---
    # Remove colunas internas de controle — mantém apenas dados do ERP + valores calculados
    COLUNAS_INTERNAS = {
        "indenizacao_ja_faturada", "despesa_ja_faturada", "incluir_indenizacao",
        "incluir_despesa", "motivo_exclusao", "usar_valor_erp_inden",
        "pct_franquia", "pct_liquido",
    }
    colunas_lote = [c for c in df_lote.columns if c not in COLUNAS_INTERNAS]
    df_export = df_lote[colunas_lote].copy()

    # Renomeia coberturas no arquivo final (De → Para)
    COBERTURA_DEPARA = {
        "Roubo e Furto": "Roubo ou Furto Qualificado",
    }
    col_cob_export = _encontrar_coluna_fuzzy(df_export, ["Cobertura de Seguro", "cobertura"])
    if col_cob_export:
        df_export[col_cob_export] = df_export[col_cob_export].replace(COBERTURA_DEPARA)

    # PERCENTUAL COBRADO FRANQUIA: substitui ponto por vírgula
    col_pct_export = _encontrar_coluna_fuzzy(df_export, ["PERCENTUAL COBRADO"])
    if col_pct_export:
        df_export[col_pct_export] = df_export[col_pct_export].astype(str).str.replace(".", ",", regex=False)

    caminho_lote = OUTPUT_DIR / f"Lote_AXA_{data_hoje}_{LOTE_ATUAL.replace(' ', '_')}.xlsx"
    with pd.ExcelWriter(caminho_lote, engine="openpyxl") as writer:
        df_export.to_excel(writer, sheet_name="Lote", index=False)
        _formatar_excel(writer, "Lote")

    print(f"[Export] Lote salvo: {caminho_lote}")

    # --- Relatório de Conferência ---
    caminho_conf = OUTPUT_DIR / f"Conferencia_AXA_{data_hoje}.xlsx"

    with pd.ExcelWriter(caminho_conf, engine="openpyxl") as writer:
        # Aba completa com todas as validações
        df_completo.to_excel(writer, sheet_name="Todos os Sinistros", index=False)

        # Aba de divergências — apenas sinistros no lote com valor diferente do ERP
        # Exclui casos onde ERP = 0 e nosso valor > 0 (intencional — usamos o calculado)
        no_lote = df_completo["incluir_indenizacao"] | df_completo["incluir_despesa"]
        col_erp_inden = _encontrar_coluna_fuzzy(df_completo, ["Valor Pagto Indeniza"])
        col_erp_desp  = _encontrar_coluna_fuzzy(df_completo, ["Valor Pagto Despesa"])
        erp_inden_zero = df_completo[col_erp_inden].fillna(0).astype(float) == 0 if col_erp_inden else False
        erp_desp_zero  = df_completo[col_erp_desp].fillna(0).astype(float) == 0  if col_erp_desp  else False

        div_inden = (df_completo["diferenca_indenizacao"].abs() > 0.01) & ~erp_inden_zero
        div_desp  = (df_completo["diferenca_despesa"].abs() > 0.01)     & ~erp_desp_zero

        divergencias = df_completo[no_lote & (div_inden | div_desp)]
        if not divergencias.empty:
            divergencias.to_excel(writer, sheet_name="Divergências", index=False)

        # Aba de excluídos — com valores já faturados do histórico
        excluidos = df_completo[~(df_completo["incluir_indenizacao"] | df_completo["incluir_despesa"])].copy()
        if not excluidos.empty:
            if df_hist is not None and not df_hist.empty:
                excluidos = _enriquecer_com_historico(excluidos, df_hist)
            excluidos.to_excel(writer, sheet_name="Excluídos", index=False)

        # Aba de resumo
        _gerar_resumo(df_completo, writer, df_hist)

    print(f"[Export] Conferência salva: {caminho_conf}")

    # --- Fatura Única ---
    caminho_fatura = OUTPUT_DIR / f"FaturaUnica_{data_hoje}_{LOTE_ATUAL.replace(' ', '_')}.xlsx"
    gerar_fatura_unica(df_lote, caminho_fatura)

    return caminho_lote, caminho_conf, caminho_fatura


# ---------------------------------------------------------------------------
# Fatura Única
# ---------------------------------------------------------------------------

_COD_TIPO_OPERACAO   = {"Indenização": 1, "Honorários": 5, "Despesas": 6}
_COD_FORMA_PAGAMENTO = {"Total": 0, "Parcial": 1}
_COD_TIPO_LIQUIDACAO = {"Crédito em Conta": 1, "Bordero": 6}
_COD_COBERTURA = {
    "Quebra Acidental": 2698,
    "Roubo e Furto Qualificado": 2805,
    "Extensão da Garantia Original": 2353,
    "Derramamento de Líquido": 2221,
    "Danos elétricos": 2198,
}
_COD_CONTATO = {"Seguradora": 42}

_COBERTURA_DEPARA_FATURA = {
    "Roubo e Furto": "Roubo e Furto Qualificado",
    "Roubo ou Furto Qualificado": "Roubo e Furto Qualificado",
}

_HEADERS_FATURA = [
    "Nr Aviso", "Nr Sinistro", "CPF/CNPJ", "Valor", "Tipo da Cobertura",
    "Tipo Operação", "Tipo Liquidação", "Forma Pagamento", "Nº ISJ",
    "Nº Nota Fiscal", "Data Emissão Nota Fiscal", "Descrição", "Ex Gratia",
    "Tipo Contato", "Cod Tipo Operação", "Cod Forma Pagamento",
    "Cod Tipo Liquidação", "Cod Cobertura", "Cod Contato",
    "Validate", "Estado Pagamento (não preencher)", "Descrição de Erros (não preencher)",
]


def gerar_fatura_unica(df_lote: pd.DataFrame, caminho: Path):
    """Gera o arquivo Fatura Única no formato exigido pela AXA (uma linha por tipo de cobrança)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    col_sinistro   = _encontrar_coluna_fuzzy(df_lote, ["Sinistro AXA"])
    col_cpf        = _encontrar_coluna_fuzzy(df_lote, ["CPF"])
    col_cobertura  = _encontrar_coluna_fuzzy(df_lote, ["Cobertura de Seguro", "cobertura"])
    col_finalizacao= _encontrar_coluna_fuzzy(df_lote, ["Tipo de Finalização", "Tipo de Finalizacao"])
    col_os         = _encontrar_coluna_fuzzy(df_lote, ["OS Interna PLL", "OS Interna"])

    def _limpar(v):
        return str(v or "").strip().strip("`")

    rows = []
    for _, row in df_lote.iterrows():
        os_interna  = _limpar(row.get(col_os, "") if col_os else "")
        sinistro    = _limpar(row.get(col_sinistro, "") if col_sinistro else "")
        cpf         = _limpar(row.get(col_cpf, "") if col_cpf else "")
        cob_raw     = _limpar(row.get(col_cobertura, "") if col_cobertura else "")
        cobertura   = _COBERTURA_DEPARA_FATURA.get(cob_raw, cob_raw)
        finalizacao = _limpar(row.get(col_finalizacao, "") if col_finalizacao else "")

        incluir_desp  = bool(row.get("incluir_despesa", False))
        incluir_inden = bool(row.get("incluir_indenizacao", False))
        val_desp      = float(row.get("valor_despesa_correto", 0) or 0)
        val_inden     = float(row.get("valor_indenizacao_correto", 0) or 0)

        cod_cob  = _COD_COBERTURA.get(cobertura, "")
        cod_cont = _COD_CONTATO.get("Seguradora", 42)
        cod_fp   = _COD_FORMA_PAGAMENTO.get("Total", 0)
        cod_liq  = _COD_TIPO_LIQUIDACAO.get("Crédito em Conta", 1)

        def _linha(tipo_op, valor, row_num):
            # Colunas de fórmula usam referências à própria linha e à aba Aux
            f_cod_op  = f'=IFERROR(VLOOKUP(F{row_num},Aux!$A$2:$B$8,2,FALSE),"")'
            f_cod_fp  = f'=IFERROR(VLOOKUP(H{row_num},Aux!$D$2:$E$3,2,FALSE),"")'
            f_cod_liq = f'=IFERROR(VLOOKUP(G{row_num},Aux!$D$8:$E$9,2,FALSE),"")'
            f_cod_cob = f'=IFERROR(VLOOKUP(E{row_num},Aux!$J$2:$K$18,2,FALSE),"")'
            f_cod_cont= f'=IFERROR(VLOOKUP(N{row_num},Aux!$S$3:$T$31,2,FALSE),"")'
            f_validate= (
                f'=IFERROR(IF(AND(IF(AND(OR(LEN(A{row_num})>0,LEN(B{row_num})>0),'
                f'OR(LEN(C{row_num})>10,LEN(C{row_num})>10),D{row_num}>0,'
                f'AND(O{row_num}>0,NOT(O{row_num}="")),AND(P{row_num}>=0,NOT(P{row_num}="")),'
                f'LEN(Q{row_num})>0),1,0)=0,NOT(LEN(O{row_num})>0),NOT(LEN(P{row_num})>0),'
                f'NOT(LEN(Q{row_num})>0)),"",IF(AND(OR(LEN(A{row_num})>0,LEN(B{row_num})>0),'
                f'LEN(C{row_num})>10,(LEN(D{row_num})>0),D{row_num}>0,(LEN(O{row_num})>0),'
                f'O{row_num}>0,(LEN(P{row_num})>0),P{row_num}>=0,LEN(Q{row_num})>0,'
                f'Q{row_num}>0,IF(AND(O{row_num}=5,NOT(LEN(J{row_num})>0)),0,1)),1,0)),"")'
            )
            return [
                os_interna, sinistro, cpf, valor, cobertura,
                tipo_op, "Crédito em Conta", "Total", None,
                None, None, finalizacao, 1,
                "Seguradora", f_cod_op, f_cod_fp, f_cod_liq, f_cod_cob, f_cod_cont,
                f_validate, None, None,
            ]

        if incluir_desp and val_desp > 0:
            rows.append(_linha("Despesas", val_desp, len(rows) + 2))
        if incluir_inden and val_inden > 0:
            rows.append(_linha("Indenização", val_inden, len(rows) + 2))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fatura Unica"

    # Cabeçalho
    ws.append(_HEADERS_FATURA)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Dados
    for r in rows:
        ws.append(r)

    # Largura automática
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 35)
    ws.row_dimensions[1].height = 30

    # Aba Aux (tabelas de referência usadas no modelo AXA)
    ws_aux = wb.create_sheet("Aux")
    _criar_aux_fatura(ws_aux)

    wb.save(str(caminho))
    print(f"[Fatura] Fatura Unica salva: {caminho} ({len(rows)} linhas)")


def _criar_aux_fatura(ws):
    """Popula a aba Aux com as tabelas de referência da Fatura Única."""
    # Tipo Operação (col A:B)
    ws["A1"] = "Transacao Financeira Tipo"
    dados_op = [("Indenização", 1), ("Honorários", 5), ("Despesas", 6)]
    for i, (nome, cod) in enumerate(dados_op, 2):
        ws.cell(row=i, column=1, value=nome)
        ws.cell(row=i, column=2, value=cod)

    # Forma Pagamento / Tipo Liquidação (col D:E)
    ws["D1"] = "Forma Pagamento"
    for i, (nome, cod) in enumerate([("Parcial", 1), ("Total", 0)], 2):
        ws.cell(row=i, column=4, value=nome)
        ws.cell(row=i, column=5, value=cod)
    ws["D7"] = "Tipo Liquidacao"
    for i, (nome, cod) in enumerate([("Crédito em Conta", 1), ("Bordero", 6)], 8):
        ws.cell(row=i, column=4, value=nome)
        ws.cell(row=i, column=5, value=cod)

    # Tipo Cobertura (col J:K)
    ws["J1"] = "Tipo Cobertura"
    coberturas = [
        ("Derramamento de Líquido", 2221), ("Extensão da Garantia Original", 2353),
        ("Quebra Acidental", 2698), ("Roubo e Furto Qualificado", 2805), ("Danos elétricos", 2198),
    ]
    for i, (nome, cod) in enumerate(coberturas, 2):
        ws.cell(row=i, column=10, value=nome)
        ws.cell(row=i, column=11, value=cod)

    # Tipo Contato (col S:T)
    ws["S2"] = "Contato Tipo"
    contatos = [
        ("Advogado", 21), ("Agente", 23), ("Assessoria", 47), ("Banco", 49),
        ("Beneficiarios", 19), ("Congenere", 44), ("Corretor", 27), ("Estipulante", 48),
        ("Fiador", 25), ("Investigadores", 14), ("Laboratorio", 24), ("Leiloeiro", 36),
        ("Medico", 43), ("Pagador", 26), ("PeritoAviacao", 32), ("PeritoConstrucao", 33),
        ("PeritoEngenharia", 30), ("PeritoOutros", 35), ("PeritoPatrimoniais", 29),
        ("PeritoRCivil", 31), ("PeritoTransportes", 34), ("Prolaborista", 28),
        ("Ressegurador", 39), ("Segurado", 7), ("Seguradora", 42),
        ("Subestipulante", 38), ("Testemunha", 5), ("Tomador", 1),
    ]
    for i, (nome, cod) in enumerate(contatos, 3):
        ws.cell(row=i, column=19, value=nome)
        ws.cell(row=i, column=20, value=cod)


def _formatar_excel(writer, sheet_name: str):
    """Aplica largura automática de colunas."""
    try:
        ws = writer.sheets[sheet_name]
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    except Exception:
        pass


def _enriquecer_com_historico(df_excluidos: pd.DataFrame, hist: pd.DataFrame) -> pd.DataFrame:
    """Adiciona colunas com os valores já faturados no histórico para cada OS excluída."""
    col_os_hist  = _encontrar_coluna_fuzzy(hist, ["OS Interna PLL", "OS Interna"])
    col_os_df    = _encontrar_coluna_fuzzy(df_excluidos, ["OS Interna PLL", "OS Interna"])
    col_i_hist   = _encontrar_coluna_fuzzy(hist, ["Valor Pagto Indeniza"])
    col_d_hist   = _encontrar_coluna_fuzzy(hist, ["Valor Pagto Despesa"])

    if not col_os_hist or not col_os_df:
        return df_excluidos

    def _v(s):
        v = str(s).replace("R$","").replace(" ","").replace(".","").replace(",",".").strip()
        try: return float(v)
        except: return 0.0

    hist = hist.copy()
    hist["_os"]    = hist[col_os_hist].astype(str).str.strip()
    hist["_inden"] = hist[col_i_hist].apply(_v) if col_i_hist else 0
    hist["_desp"]  = hist[col_d_hist].apply(_v) if col_d_hist else 0

    agg = hist.groupby("_os").agg(
        hist_inden=("_inden", "sum"),
        hist_desp=("_desp",  "sum"),
    ).reset_index().rename(columns={"_os": col_os_df})

    df_excluidos = df_excluidos.merge(agg, on=col_os_df, how="left")
    df_excluidos["hist_inden"] = df_excluidos["hist_inden"].fillna(0)
    df_excluidos["hist_desp"]  = df_excluidos["hist_desp"].fillna(0)
    df_excluidos = df_excluidos.rename(columns={
        "hist_inden": "Hist: Indenizacao ja faturada (R$)",
        "hist_desp":  "Hist: Despesa ja faturada (R$)",
    })
    return df_excluidos


def _gerar_resumo(df: pd.DataFrame, writer, df_hist: pd.DataFrame = None):
    """Gera aba de resumo com totais e detalhamento dos motivos de exclusão."""
    lote      = df[df["incluir_indenizacao"] | df["incluir_despesa"]]
    excluidos = df[~(df["incluir_indenizacao"] | df["incluir_despesa"])]

    total_indenizacao = lote.loc[lote["incluir_indenizacao"], "valor_indenizacao_correto"].sum()
    total_despesa     = lote.loc[lote["incluir_despesa"],     "valor_despesa_correto"].sum()
    total_geral       = total_indenizacao + total_despesa

    resumo = pd.DataFrame({
        "Indicador": [
            "Total de sinistros processados",
            "Sinistros no lote",
            "Sinistros excluidos",
            "Total Indenizacao (R$)",
            "Total Despesa (R$)",
            "Total Geral (R$)",
            "Lote",
            "Data de geracao",
        ],
        "Valor": [
            len(df),
            len(lote),
            len(excluidos),
            f"R$ {total_indenizacao:,.2f}",
            f"R$ {total_despesa:,.2f}",
            f"R$ {total_geral:,.2f}",
            LOTE_ATUAL,
            datetime.today().strftime("%d/%m/%Y %H:%M"),
        ],
    })
    resumo.to_excel(writer, sheet_name="Resumo", index=False, startrow=0)

    try:
        from openpyxl.styles import Font, PatternFill, Alignment

        def _bloco(writer, sheet, titulo, df_bloco, startrow, cor="1F4E79"):
            ws = writer.sheets[sheet]
            # Título do bloco
            ws.cell(row=startrow, column=1, value=titulo).font = Font(bold=True, color="FFFFFF", size=11)
            ws.cell(row=startrow, column=1).fill = PatternFill("solid", fgColor=cor)
            ws.cell(row=startrow, column=2).fill = PatternFill("solid", fgColor=cor)
            startrow += 1
            # Cabeçalho
            for ci, col in enumerate(df_bloco.columns, 1):
                c = ws.cell(row=startrow, column=ci, value=col)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="2E75B6")
            startrow += 1
            # Dados
            for _, row in df_bloco.iterrows():
                for ci, val in enumerate(row, 1):
                    ws.cell(row=startrow, column=ci, value=val)
                startrow += 1
            return startrow + 1

        ws = writer.sheets["Resumo"]
        ws.column_dimensions["A"].width = 85
        ws.column_dimensions["B"].width = 15

        linha = len(resumo) + 3

        # Bloco 1: Motivos de exclusão + totais do histórico
        if not excluidos.empty:
            motivos = excluidos["motivo_exclusao"].value_counts(dropna=False).reset_index()
            motivos.columns = ["Motivo de Exclusao", "Quantidade"]

            # Totais já faturados no histórico para os excluídos
            if df_hist is not None and not df_hist.empty:
                enriquecido = _enriquecer_com_historico(excluidos, df_hist)
                total_hist_inden = enriquecido["Hist: Indenizacao ja faturada (R$)"].sum()
                total_hist_desp  = enriquecido["Hist: Despesa ja faturada (R$)"].sum()
                motivos = pd.concat([motivos, pd.DataFrame([
                    {"Motivo de Exclusao": "─" * 40, "Quantidade": ""},
                    {"Motivo de Exclusao": "Total ja faturado no historico - Indenizacao", "Quantidade": f"R$ {total_hist_inden:,.2f}"},
                    {"Motivo de Exclusao": "Total ja faturado no historico - Despesa",     "Quantidade": f"R$ {total_hist_desp:,.2f}"},
                    {"Motivo de Exclusao": "TOTAL JA FATURADO",                             "Quantidade": f"R$ {total_hist_inden + total_hist_desp:,.2f}"},
                ])], ignore_index=True)

            linha = _bloco(writer, "Resumo", "DETALHAMENTO DOS EXCLUIDOS", motivos, linha)

        # Bloco 2: Divergências de valor (apenas no lote)
        no_lote = df[df["incluir_indenizacao"] | df["incluir_despesa"]]
        col_ei = next((c for c in df.columns if "Valor Pagto Indeniza" in c), None)
        col_ed = next((c for c in df.columns if "Valor Pagto Despesa"  in c), None)
        erp_i_zero = df[col_ei].fillna(0).astype(float) == 0 if col_ei else pd.Series(False, index=df.index)
        erp_d_zero = df[col_ed].fillna(0).astype(float) == 0 if col_ed else pd.Series(False, index=df.index)
        no_lote_mask = df["incluir_indenizacao"] | df["incluir_despesa"]
        mask_div = no_lote_mask & (
            ((df["diferenca_indenizacao"].abs() > 0.01) & ~erp_i_zero) |
            ((df["diferenca_despesa"].abs()     > 0.01) & ~erp_d_zero)
        )
        divs = df[mask_div]
        if not divs.empty:
            div_resumo = divs[["regra_aplicada", "diferenca_indenizacao", "diferenca_despesa"]].copy()
            div_resumo.columns = ["Regra Aplicada", "Dif. Indenizacao (R$)", "Dif. Despesa (R$)"]
            div_resumo = div_resumo.groupby("Regra Aplicada").agg(
                Quantidade=("Dif. Indenizacao (R$)", "count"),
                Total_Dif_Inden=("Dif. Indenizacao (R$)", "sum"),
                Total_Dif_Desp=("Dif. Despesa (R$)", "sum"),
            ).reset_index()
            div_resumo.columns = ["Regra Aplicada", "Quantidade", "Total Dif. Indenizacao", "Total Dif. Despesa"]
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 22
            _bloco(writer, "Resumo", "DETALHAMENTO DAS DIVERGENCIAS", div_resumo, linha, cor="833C00")

    except Exception as e:
        print(f"[Resumo] Aviso formatacao: {e}")
